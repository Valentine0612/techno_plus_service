from openpyxl import load_workbook
import os
import json


path = os.path.join(os.getcwd(), 'guide/utils/excel_db/source')
file = os.path.join(path, 'guide.xlsx')

wb = load_workbook(file)

ws = wb['11. Справ. инстр.оснаст.приспос']

data = []
ID_CHAPTER = 0
ID_GROUP = 0
ID_RESOURCE = 0
MISSED = 0



for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H']):
    if MISSED < 5:
        MISSED += 1
        continue
    else:
        if el[0].value:
            ID_CHAPTER += 1
            chapter_data = {
                'model': 'snap_in_tool.chapter',
                'pk': ID_CHAPTER,
                'fields' : {
                        'title': el[0].value,
                        'description': el[1].value,
                }
            }
            data.append(chapter_data)
        if el[2].value:
            ID_GROUP += 1
            group_data = {
                'model': 'snap_in_tool.Group',
                'pk': ID_GROUP,
                'fields' : {
                        'title': el[2].value,
                        'description': el[3].value,
                        'chapter' : ID_CHAPTER,
                }
            }
            data.append(group_data)
        if el[4].value:
            ID_RESOURCE += 1
            resource_data = {
                'model': 'snap_in_tool.Resource',
                'pk': ID_RESOURCE,
                'fields' : {
                        'title': el[4].value,
                        'description': el[5].value,
                        'measurement' : el[6].value,
                        'group': ID_GROUP,
                }
            }
            data.append(resource_data)   


path = os.path.join(os.getcwd(), 'guide/snap_in_tool/fixtures')
file = os.path.join(path, 'snap_in_tool.json')
with open(file, 'w', encoding='utf-8') as f:
    json.dump(data, f)
