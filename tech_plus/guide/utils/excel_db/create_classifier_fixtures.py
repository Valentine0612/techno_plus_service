from openpyxl import load_workbook
import os
import json


path = os.path.join(os.getcwd(), 'guide/utils/excel_db/source')
file = os.path.join(path, 'БД_справочников_ТЕХНО_10_06_2021_Классификатор_ТК_ГЭСН,_КСР_zgbFlEk.xlsx')

wb = load_workbook(file)

ws = wb['9. Классификатор КСР']

data = []
ID_BOOK = 0
ID_PARTBOOK = 0
ID_CHAPTER = 0
ID_GROUP = 0
ID_RESOURCE = 0

MISSED = 0


for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['J'], ws['K'], ws['L']):
    if MISSED < 4:
        MISSED += 1
        continue
    else:
        if el[1].value:
            ID_BOOK += 1
            book_data = {
                'model': 'classifier_kcp.book',
                'pk': ID_BOOK,
                'fields' : {
                        'title': el[0].value,
                        'description': el[1].value,
                }
            }
            data.append(book_data)

        if el[2].value:
            ID_PARTBOOK += 1
            part_book_data = {
                'model': 'classifier_kcp.PartBook',
                'pk': ID_PARTBOOK,
                'fields' : {
                        'title': el[2].value,
                        'description': el[3].value,
                        'book' : ID_BOOK,
                }
            }
            data.append(part_book_data)

        if el[4].value:
            ID_CHAPTER += 1
            chapter_data = {
                'model': 'classifier_kcp.Chapter',
                'pk': ID_CHAPTER,
                'fields' : {
                        'title': el[4].value,
                        'description': el[5].value,
                        'partbook' : ID_PARTBOOK,
                }
            }
            data.append(chapter_data)

        if el[6].value:
            ID_GROUP += 1
            group_data = {
                'model': 'classifier_kcp.Group',
                'pk': ID_GROUP,
                'fields' : {
                        'title': el[6].value,
                        'description': el[7].value,
                        'chapter' : ID_CHAPTER,
                }
            }
            data.append(group_data)

        if el[8].value:
            ID_RESOURCE += 1
            resource_data = {
                'model': 'classifier_kcp.Resource',
                'pk': ID_RESOURCE,
                'fields' : {
                        'title': el[8].value,
                        'description': el[9].value,
                        'measurement' : el[10].value,
                        'group': ID_GROUP,
                }
            }
            data.append(resource_data)   

path = os.path.join(os.getcwd(), 'guide/classifier_kcp/fixtures')
file = os.path.join(path, 'classifier.json')
with open(file, 'w', encoding='utf-8') as f:
    json.dump(data, f)
