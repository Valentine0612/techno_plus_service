import abc
import os
import json
from openpyxl import load_workbook


class CreateTk(abc.ABC):
    ''' класс отвечает за создание excel  '''

    @abc.abstractclassmethod
    def get_data():
        pass

    @staticmethod
    def create_db(wb, type):
        TYPE = {
            'TkMain': TkMain,
            'TkRemoteConstruction': TkRemoteConstruction,
            'TkInstallationEquipment': TkInstallationEquipment,
            'TkPre_commissioning': TkPre_commissioning,
        }
        return TYPE[type].get_data(wb)


class TkMain(CreateTk):

    def get_data(wb):
        sheet = '2. ТК ГЭСН'
        name = 'tk_main'
        data = []


        ws = wb[sheet]
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_SUBSECTION = 0
        ID_TABLESUBSECTION = 0
        ID_TECHCARD = 0
        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_SUBSECTION = False
        ACCEPT_TABLESUBSECTION = False
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K'], ws['L'], ws['M']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                if el[0].value:
                    ID_COLLECTION += 1
                    collection_data = {
                        'model': 'tk_gesn.TkMainCollection',
                        'pk': ID_COLLECTION,
                        'fields' : {
                                'title': el[0].value,
                                'description': el[1].value if el[1].value else 'None',
                        }
                    }
                    data.append(collection_data)
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                

                if el[2].value:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkMainDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': el[2].value,
                                'description': el[3].value if el[3].value else 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkMainDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    


                if el[4].value:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkMainSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': el[4].value,
                                'description': el[5].value if el[5].value else 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkMainSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_SECTION = False


                if el[6].value:
                    ID_SUBSECTION += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkMainSubsection',
                        'pk': ID_SUBSECTION,
                        'fields' : {
                                'title': el[6].value,
                                'description': el[7].value if el[7].value else 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SUBSECTION:
                    ID_SUBSECTION += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkMainSubsection',
                        'pk': ID_SUBSECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_SUBSECTION = False

                if el[8].value:
                    ID_TABLESUBSECTION += 1
                    table_subsection_data = {
                        'model': 'tk_gesn.TkMainTableSubsection',
                        'pk': ID_TABLESUBSECTION,
                        'fields' : {
                                'title': el[8].value,
                                'description': el[9].value if el[9].value else 'None',
                                'subsection': ID_SUBSECTION,
                        }
                    }
                    data.append(table_subsection_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = False
                    continue
                elif ACCEPT_TABLESUBSECTION:
                    ID_TABLESUBSECTION += 1
                    table_subsection_data = {
                        'model': 'tk_gesn.TkMainTableSubsection',
                        'pk': ID_TABLESUBSECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'subsection': ID_SUBSECTION,
                        }
                    }
                    data.append(table_subsection_data)
                    ACCEPT_TABLESUBSECTION = False
                    
                         
                if el[10].value:
                    ID_TECHCARD += 1
                    techcard_data = {
                        'model': 'tk_gesn.TkMainTechCard',
                        'pk': ID_TECHCARD,
                        'fields' : {
                                'title': el[10].value,
                                'description': el[11].value if el[11].value else 'None',
                                'table_subsection': ID_TABLESUBSECTION,
                        }
                    }
                    data.append(techcard_data)      

        path = os.path.join(os.getcwd(), f'guide/tk_gesn/fixtures')
        try:
            os.mkdir(path)
        except OSError:
            pass
        file = os.path.join(path, f'{name}.json')
        with open(file, 'w', encoding='utf-8') as f:
            json.dump(data, f)


class TkRemoteConstruction(CreateTk):

    def get_data(wb):
        sheet = '2.1. ТК ГЭСН Рем.-строит работ'
        name = 'tk_remote_construction'
        data = []

        ws = wb[sheet]
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_SUBSECTION = 0
        ID_TECHCARD = 0
        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_SUBSECTION = False
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                if el[0].value:
                    ID_COLLECTION += 1
                    collection_data = {
                        'model': 'tk_gesn.TkRemoteConstructionCollection',
                        'pk': ID_COLLECTION,
                        'fields' : {
                                'title': el[0].value,
                                'description': el[1].value if el[1].value else 'None',
                        }
                    }
                    data.append(collection_data)
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    continue
                

                if el[2].value:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkRemoteConstructionDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': el[2].value,
                                'description': el[3].value if el[3].value else 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkRemoteConstructionDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    


                if el[4].value:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkRemoteConstructionSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': el[4].value,
                                'description': el[5].value if el[5].value else 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = True
                    
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkRemoteConstructionSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_SECTION = False


                if el[6].value:
                    ID_SUBSECTION += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkRemoteConstructionSubsection',
                        'pk': ID_SUBSECTION,
                        'fields' : {
                                'title': el[6].value,
                                'description': el[7].value if el[7].value else 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    continue
                elif ACCEPT_SUBSECTION:
                    ID_SUBSECTION += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkRemoteConstructionSubsection',
                        'pk': ID_SUBSECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_SUBSECTION = False
                    
                         
                if el[8].value:
                    ID_TECHCARD += 1
                    techcard_data = {
                        'model': 'tk_gesn.TkRemoteConstructionTechCard',
                        'pk': ID_TECHCARD,
                        'fields' : {
                                'title': el[8].value,
                                'description': el[9].value if el[9].value else 'None',
                                'subsection': ID_SUBSECTION,
                        }
                    }
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    data.append(techcard_data)      

        path = os.path.join(os.getcwd(), f'guide/tk_gesn/fixtures')
        try:
            os.mkdir(path)
        except OSError:
            pass
        file = os.path.join(path, f'{name}.json')
        with open(file, 'w', encoding='utf-8') as f:
            json.dump(data, f)


class TkInstallationEquipment(CreateTk):

    def get_data(wb):
        sheet = '2.2. ТК ГЭСН Монтаж оборуд.'
        name = 'tk_installation_equipment'
        data = []

        ws = wb[sheet]
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_TABLE = 0
        ID_TECHCARD = 0
        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_TABLE = False
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                if el[0].value:
                    ID_COLLECTION += 1
                    collection_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentCollection',
                        'pk': ID_COLLECTION,
                        'fields' : {
                                'title': el[0].value,
                                'description': el[1].value if el[1].value else 'None',
                        }
                    }
                    data.append(collection_data)
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_TABLE = True
                    continue
                

                if el[2].value:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': el[2].value,
                                'description': el[3].value if el[3].value else 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_TABLE = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    


                if el[4].value:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': el[4].value,
                                'description': el[5].value if el[5].value else 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_TABLE = True
                    
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_SECTION = False


                if el[6].value:
                    ID_TABLE += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentTable',
                        'pk': ID_TABLE,
                        'fields' : {
                                'title': el[6].value,
                                'description': el[7].value if el[7].value else 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_TABLE = False
                    continue
                elif ACCEPT_TABLE:
                    ID_TABLE += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentTable',
                        'pk': ID_TABLE,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_TABLE = False
                    
                         
                if el[8].value:
                    ID_TECHCARD += 1
                    techcard_data = {
                        'model': 'tk_gesn.TkInstallationEquipmentTechCard',
                        'pk': ID_TECHCARD,
                        'fields' : {
                                'title': el[8].value,
                                'description': el[9].value if el[9].value else 'None',
                                'table': ID_TABLE,
                        }
                    }
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_TABLE = False
                    data.append(techcard_data)      

        path = os.path.join(os.getcwd(), f'guide/tk_gesn/fixtures')
        try:
            os.mkdir(path)
        except OSError:
            pass
        file = os.path.join(path, f'{name}.json')
        with open(file, 'w', encoding='utf-8') as f:
            json.dump(data, f)


class TkPre_commissioning(CreateTk):

    def get_data(wb):
        sheet = '2.3. ТК  ГЭСН Пусконаладочные'
        name = 'tk_pre_commissioning'
        data = []

        ws = wb[sheet]
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_SUBSECTION = 0
        ID_TABLESUBSECTION = 0
        ID_TECHCARD = 0
        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_SUBSECTION = False
        ACCEPT_TABLESUBSECTION = False
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K'], ws['L'], ws['M']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                if el[0].value:
                    ID_COLLECTION += 1
                    collection_data = {
                        'model': 'tk_gesn.TkPre_commissioningCollection',
                        'pk': ID_COLLECTION,
                        'fields' : {
                                'title': el[0].value,
                                'description': el[1].value if el[1].value else 'None',
                        }
                    }
                    data.append(collection_data)
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                

                if el[2].value:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkPre_commissioningDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': el[2].value,
                                'description': el[3].value if el[3].value else 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT += 1
                    departmnet_data = {
                        'model': 'tk_gesn.TkPre_commissioningDepartment',
                        'pk': ID_DEPARTMENT,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'collection' : ID_COLLECTION,
                        }
                    }
                    data.append(departmnet_data)
                    ACCEPT_DEPARTMENT = False
                    


                if el[4].value:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkPre_commissioningSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': el[4].value,
                                'description': el[5].value if el[5].value else 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION += 1
                    section_data = {
                        'model': 'tk_gesn.TkPre_commissioningSection',
                        'pk': ID_SECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'department': ID_DEPARTMENT,
                        }
                    }
                    data.append(section_data)
                    ACCEPT_SECTION = False


                if el[6].value:
                    ID_SUBSECTION += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkPre_commissioningSubsection',
                        'pk': ID_SUBSECTION,
                        'fields' : {
                                'title': el[6].value,
                                'description': el[7].value if el[7].value else 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SUBSECTION:
                    ID_SUBSECTION += 1
                    subsection_data = {
                        'model': 'tk_gesn.TkPre_commissioningSubsection',
                        'pk': ID_SUBSECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'section': ID_SECTION,
                        }
                    }
                    data.append(subsection_data)
                    ACCEPT_SUBSECTION = False

                if el[8].value:
                    ID_TABLESUBSECTION += 1
                    table_subsection_data = {
                        'model': 'tk_gesn.TkPre_commissioningTableSubsection',
                        'pk': ID_TABLESUBSECTION,
                        'fields' : {
                                'title': el[8].value,
                                'description': el[9].value if el[9].value else 'None',
                                'subsection': ID_SUBSECTION,
                        }
                    }
                    data.append(table_subsection_data)
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = False
                    continue
                elif ACCEPT_TABLESUBSECTION:
                    ID_TABLESUBSECTION += 1
                    table_subsection_data = {
                        'model': 'tk_gesn.TkPre_commissioningTableSubsection',
                        'pk': ID_TABLESUBSECTION,
                        'fields' : {
                                'title': 'None',
                                'description': 'None',
                                'subsection': ID_SUBSECTION,
                        }
                    }
                    data.append(table_subsection_data)
                    ACCEPT_TABLESUBSECTION = False
                    
                         
                if el[10].value:
                    ID_TECHCARD += 1
                    techcard_data = {
                        'model': 'tk_gesn.TkPre_commissioningTechCard',
                        'pk': ID_TECHCARD,
                        'fields' : {
                                'title': el[10].value,
                                'description': el[11].value if el[11].value else 'None',
                                'table_subsection': ID_TABLESUBSECTION,
                        }
                    }
                    data.append(techcard_data)      

        path = os.path.join(os.getcwd(), f'guide/tk_gesn/fixtures')
        try:
            os.mkdir(path)
        except OSError:
            pass
        file = os.path.join(path, f'{name}.json')
        with open(file, 'w', encoding='utf-8') as f:
            json.dump(data, f)


path = os.path.join(os.getcwd(), 'guide/utils/excel_db/source')
file = os.path.join(path, 'guide.xlsx')

wb = load_workbook(file)
CreateTk.create_db(wb, 'TkMain')
CreateTk.create_db(wb, 'TkRemoteConstruction')
CreateTk.create_db(wb, 'TkInstallationEquipment')
CreateTk.create_db(wb, 'TkPre_commissioning')