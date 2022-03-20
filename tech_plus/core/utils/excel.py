import abc
import os
import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill
from openpyxl.cell import Cell
import datetime
import json

from core.models import CheckExcel



class GetId():

    @staticmethod
    def new(model_obj):
        return model_obj.objects.all().count() + 1

    @staticmethod
    def last(model_obj):
        return model_obj.objects.all().count()

    @staticmethod
    def current(model_obj, pk):
        return model_obj.objects.all().count()

def end_document(ws, count, start_row='b'):
    start_row = start_row.upper()
    ft = Font(name='Times New Roman')
    ws[start_row+str(count+2)] = 'Составил _________________________________'
    ws[start_row+str(count+2)].font = ft

    ws[start_row+str(count+3)] = '    (должность, подпись, инициалы, фамилия)'
    ws[start_row+str(count+3)].font = ft

    ws[start_row+str(count+5)] = 'Проверил _________________________________'
    ws[start_row+str(count+5)].font = ft

    ws[start_row+str(count+6)] = '     (должность, подпись, инициалы, фамилия)'
    ws[start_row+str(count+6)].font = ft


def sum_colls(arr):
            answer = 0
            for el in arr:
                try:
                    answer += el
                except TypeError:
                    pass
            return answer

def set_border(ws, cell_range, type):
            ft = Font(name='Times New Roman')
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            rows = ws[cell_range]
            for i in rows:
                for j in i:
                    if type in ['form_1', 'form_2', 'form_3', 'form_4', 'form_5', 'form_6']:
                        j.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        j.alignment = Alignment(horizontal='center', vertical='center')
                    j.border = thin_border
                    j.font = ft

class Excel(abc.ABC):
    ''' класс отвечает за создание excel  '''

    @abc.abstractclassmethod
    def get_data(data):
        pass

    @staticmethod
    def create_excel(data):
        excel_type = data['metadata']
        TYPE = {
            'form_8': Form_8,
            'form_7': Form_7,
            'form_1': Form_1,
            'form_2': Form_2,
            'form_3': Form_3,
            'form_4': Form_4,
            'form_5': Form_5,
            'form_6': Form_6,
        }
        return TYPE[excel_type].get_data(data)

    @staticmethod
    def fill_excel(file):
        wb = load_workbook(file)
        excel_lists = wb.sheetnames
        TYPE = {
        '9. Классификатор КСР': ClassifierKcp,
        #'8. Классиф.  машины и механизмы': MachinesMechanisms,
        #'11. Справ. инстр.оснаст.приспос': SnapInTool,
        #'2. ТК ГЭСН ': TkMain, #!
        #'2.1. ТК ГЭСН Рем.-строит работ': TkRemoteConstruction,
        #'2.2. ТК ГЭСН Монтаж оборуд.': TkInstallationEquipment,
        #'2.3. ТК  ГЭСН Пусконаладочные': TkPre_commissioning,
                    }
        for excel_el in excel_lists:
            ws = wb[excel_el]
            if excel_el not in TYPE:
                continue
            else:
                print(excel_el)
                TYPE[excel_el].get_data(ws)
        check = CheckExcel.objects.get(pk=1)
        check.status = True
        check.save()


class Form_8(Excel):

    def get_data(data):

        def filter(objects, name):
            return objects.filter(name=name)

        path = './core/static/excel_template'
        file = os.path.join(path, 'form_8_example.xlsx')
        COUNT_STATIC = 16
        COUNT_OPER = 16
        COUNT_DUR = 16
        COUNT_PRODUCTS = 17
        COUNT_WORK = 18
        COUNT_PRODUCT = 19
        COUNT_NTZ = 20
        COUNT_AVG = 19
        ft = Font(name='Times New Roman')

        wb = load_workbook(file)
        ws = wb.active

            
        

        #title
        ws['F8'] = data['object'][0].title
        ws['N8'] = f"Объем выполненных работ: \n{data['object'][0].process_meter} - {data['object'][0].process_measure}"

        #date
        date = data['queryset'][0].created_at.strftime("%d/%m/%Y")
        ws['N7'] = f'Дата наблюдения {date}'

        #row C and F
        for i, oper in enumerate(data['operations']):
            ws['C'+str(COUNT_OPER)] = oper.name
            ws['F'+str(COUNT_OPER)] = oper.measure.code
            COUNT_OPER += 6

        #for ntz 
        for ntz in zip(data['ntz'].values(), data['count_of_product'].items(), data['queryset']):
            summ = sum([float(i) for i in ntz[1][1] if type(i) == float or i.isdigit()])
            items = len(ntz[1][1]) - ntz[1][1].count('-')
            ws['K' + str(COUNT_NTZ)] = f'{round(float(summ/items), 4)}×{100}'
            operation = ntz[2].has_operations.get(name=ntz[1][0])
            ws['K' + str(COUNT_NTZ+1)] = f'[100-({operation.ratio.npzr}+{operation.ratio.no}+{operation.ntp.ntp})]×60'
            ws['G'+str(COUNT_NTZ)] = f"Нтз (1 чел.)"
            ws['S'+str(COUNT_NTZ)] = ntz[0]
            COUNT_NTZ += 6

        #for views 16 string excel
        for oper_name in data['operations'].values():
            sum_string = 0
            for num_work in zip(data['queryset'], ['K', 'L', 'M', 'N', 'O']):
                obj = filter(num_work[0].has_operations, oper_name['name'])[0]
                ws[num_work[1]+str(COUNT_DUR)] = obj.duration if obj.duration != 0 else '-'
                sum_string += obj.duration
            ws['P' + str(COUNT_DUR)] = sum_string
            COUNT_DUR += 6

        #for views 17 string excel
        for oper_name in data['operations'].values():
            sum_products = 0
            for num_work in zip(data['queryset'], ['K', 'L', 'M', 'N', 'O']):
                obj = filter(num_work[0].has_operations, oper_name['name'])[0]
                ws[num_work[1]+str(COUNT_PRODUCTS)] = obj.products if obj.products != 0 else '-'
                sum_products += obj.products
            ws['P' + str(COUNT_PRODUCTS)] = sum_products
            COUNT_PRODUCTS += 6

        #for views 18 string excel
        for num_works in data['count_of_work'].values():            
            for num_work in zip(num_works, ['K', 'L', 'M', 'N', 'O']):
                ws[num_work[1]+str(COUNT_WORK)] = num_work[0] if num_work[0] != 0 else '-'
            ws['P' + str(COUNT_WORK)] = sum_colls(num_works)
            COUNT_WORK += 6

        #for views 19 string excel
        for num_works in data['count_of_product'].values():            
            for num_work in zip(num_works, ['K', 'L', 'M', 'N', 'O']):
                ws[num_work[1]+str(COUNT_PRODUCT)] = num_work[0] if num_work[0] != 0 else '-'
            ws['P' + str(COUNT_PRODUCT)] = sum_colls(num_works)
            COUNT_PRODUCT += 6
                
        #for row R
        for avg in data['average_count_of_product'].values():
            ws['R'+str(COUNT_AVG)] = avg
            COUNT_AVG += 6

            
        #for static and static excel
        for i in range(1, len(data['operations'])+1):
            ws['B'+str(COUNT_STATIC)] = f'{i}.'
            ws.merge_cells(f'B{str(COUNT_STATIC)}:B{str(COUNT_STATIC+5)}')

            ws.merge_cells(f'C{str(COUNT_STATIC)}:E{str(COUNT_STATIC+5)}')

            ws.merge_cells(f'F{str(COUNT_STATIC)}:F{str(COUNT_STATIC+5)}')


            ws['G'+str(COUNT_STATIC)] = 'Затраты труда в чел.мин'
            ws.merge_cells(f'G{str(COUNT_STATIC)}:J{str(COUNT_STATIC)}')
            ws.merge_cells(f'P{str(COUNT_STATIC)}:Q{str(COUNT_STATIC)}')

            ws['G'+str(COUNT_STATIC+1)] = 'Объем выполненных работ на измеритель'
            ws.merge_cells(f'G{str(COUNT_STATIC+1)}:J{str(COUNT_STATIC+1)}')
            ws.merge_cells(f'P{str(COUNT_STATIC+1)}:Q{str(COUNT_STATIC+1)}')

            ws['G'+str(COUNT_STATIC+2)] = 'Кол-во работ, приходящихся на 60 чел.-мин.'
            ws.merge_cells(f'G{str(COUNT_STATIC+2)}:J{str(COUNT_STATIC+2)}')
            ws.merge_cells(f'P{str(COUNT_STATIC+2)}:Q{str(COUNT_STATIC+2)}')

            ws['G'+str(COUNT_STATIC+3)] = 'Затраты на измеритель элемента, чел.-мин.'
            ws.merge_cells(f'G{str(COUNT_STATIC+3)}:J{str(COUNT_STATIC+3)}')
            ws.merge_cells(f'P{str(COUNT_STATIC+3)}:Q{str(COUNT_STATIC+3)}')

            ws.merge_cells(f'G{str(COUNT_STATIC+4)}:J{str(COUNT_STATIC+5)}')
            ws.merge_cells(f'K{str(COUNT_STATIC+4)}:R{str(COUNT_STATIC+4)}')
            ws.merge_cells(f'K{str(COUNT_STATIC+5)}:R{str(COUNT_STATIC+5)}')
            ws.merge_cells(f'S{str(COUNT_STATIC+4)}:S{str(COUNT_STATIC+5)}')
            set_border(ws, f'B{str(COUNT_STATIC)}:S{str(COUNT_STATIC+5)}', data['metadata'])

            
            COUNT_STATIC += 6


        end_document(ws=ws, count=COUNT_NTZ)
        return wb



class Form_7(Excel):

    def get_data(data):
    
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_7_example.xlsx')
        ft = Font(name='Times New Roman')
        COUNT_END = 12

        wb = load_workbook(file)
        ws = wb.active
        #for static and static excel
        for i, num_static in enumerate(range(1, len(data['object'][0].has_operations.all())+1), start=12):
            ws['B'+str(i)] = f'{num_static}.'
            ws.merge_cells(f'C{str(i)}:F{str(i)}')

            ws.merge_cells(f'BP{str(i)}:BQ{str(i)}')
            ws.row_dimensions[i].height = 35
            set_border(ws, f'B{str(i)}:BR{str(i)}', data['metadata'])
        #fill
        
        #title
        ws['B7'] = f"Наименование процесса {data['object'][0].form_8.title}"
        ws['AF5'] = f"{data['object'][0].created_at.strftime('%Y')} г."
        for i, obj in enumerate(data['object'][0].has_operations.all(), start=12):
            #color
            for el in range(obj.duration):
                start = obj.start if obj.start != 0 else 1
                color = '82b5e8'
                color_cell = wb.active.cell(column=start+el+6, row=i)
                color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            ws['BO'+str(i)] = obj.duration

            ws['BP'+str(i)] = f'{obj.products} {obj.measure.code}'

            ws['C'+str(i)] = obj.name

            workmans =''
            height = 10
            for workman in obj.member_operation.all():
                workmans += f'{workman.workman.name} - {workman.count}\n'
                height += 10
            ws['BR'+str(i)] = workmans
            ws.row_dimensions[i].height = height if height >= 40 else 40
            COUNT_END += 1

        end_document(ws=ws, count=COUNT_END)
        return wb


class Form_1(Excel):
    def get_data(data):
    
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_1_example.xlsx')

        wb = load_workbook(file)
        ws = wb.active
        COUNT_OPERATION = 13
        COUNT_STATIC = 13
        COUNT = 14
        MAX_WIDTH = 0

        #title
        ws['C6'] = data['object'][0].title
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'C6:E6')
        ws['B9'] = f"Измеритель процесса: {data['object'][0].process_meter} - {data['object'][0].process_measure}"


        #for 13 string
        for i, el in enumerate(zip(data['operations'], data['products'].values(), data['sum_products'].values()), 1):
            ws['B'+str(COUNT_OPERATION)] = f'{i}.'
            ws['C'+ str(COUNT_OPERATION)] = el[0].name
            ws['D'+ str(COUNT_OPERATION)] = el[0].measure.code
            ws['E'+ str(COUNT_OPERATION)] = f'"{el[1]}"'
            ws['F'+ str(COUNT_OPERATION)] = el[2]

            #for 15 string, materials
            ws['C'+str(COUNT)] = 'Материалы:'
            COUNT += 1
            if el[0].material_operation.all():
                for material in el[0].material_operation.all():
                    ws['C'+ str(COUNT)] = material.material.description
                    if len(material.material.description) >= MAX_WIDTH:
                        MAX_WIDTH = len(material.material.description)
                    ws['D'+ str(COUNT)] = material.material.measurement
                    ws['F'+ str(COUNT)] = material.count
                    COUNT += 1
            else:
                ws['C'+ str(COUNT)] = '...'
                COUNT += 1

            #for 17 string, machines
            ws['C'+str(COUNT)] = 'Машины и механизмы:'
            COUNT += 1
            if el[0].machine.all():
                for machine in el[0].machine.all():
                    ws['C'+ str(COUNT)] = machine.description
                    if len(machine.description) >= MAX_WIDTH:
                        MAX_WIDTH = len(machine.description)
                    ws.column_dimensions['C'].width = len(machine.description)
                    ws['D'+ str(COUNT)] = machine.measurement
                    COUNT += 1
            else:
                ws['C'+ str(COUNT)] = '...'
                COUNT += 1
            COUNT_OPERATION = COUNT
            COUNT += 1

        #for static and static excel
        COUNT_STATIC_END = COUNT_OPERATION
        set_border(ws, f'B{str(COUNT_STATIC)}:F{str(COUNT_STATIC_END-1)}', data['metadata'])

        end_document(ws=ws, count=COUNT_OPERATION)
        ws.column_dimensions['C'].width = MAX_WIDTH
        return wb



class Form_2(Excel):
    def get_data(data):
    
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_2_example.xlsx')
        ft = Font(name='Times New Roman')

        wb = load_workbook(file)
        ws = wb.active
        COUNT_STATIC = 11
        COUNT_OPERATION = 11
        COUNT = 12
        MAX_WIDTH = 0
        
        #title
        ws['D4'] = data['object'][0].title
        ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'D4:G4')
        ws['B7'] = f"Измеритель процесса: {data['object'][0].process_meter} - {data['object'][0].process_measure}"

        #for 11 string
        for i, el in enumerate(zip(data['operations'], data['sum_products'].values()), 1):
            ws['B'+str(COUNT_OPERATION)] = f'{i}.'
            ws['C'+str(COUNT_OPERATION)] = f'ТВН п. {i}'
            ws['D'+ str(COUNT_OPERATION)] = el[0].name
            ws['F'+ str(COUNT_OPERATION)] = el[0].measure.code
            ws['G'+ str(COUNT_OPERATION)] = el[1]

            #for 12 string, workman
            ws['D'+str(COUNT)] = 'Состав звена:'
            COUNT += 1
            if el[0].member_operation.all():    
                for workman in el[0].member_operation.all():
                    ws['D'+ str(COUNT)] = workman.workman.name
                    ws['E'+ str(COUNT)] = workman.count
                    ws['F'+ str(COUNT)] = workman.workman.measure
                    ws['H'+ str(COUNT)] = data['ntz'][el[0].name]
                    ws['I'+ str(COUNT)] = data['total_ntz'][el[0].name]
                    COUNT += 1
            else:
                ws['D'+ str(COUNT)] = '...'
                COUNT += 1

            #for 14 string, machines
            ws['D'+str(COUNT)] = 'Машины и механизмы:'
            COUNT += 1
            if el[0].machine_operation.all():
                for machine in el[0].machine_operation.all():
                    ws['D'+ str(COUNT)] = machine.machine.description
                    if len(machine.machine.description) >= MAX_WIDTH:
                        MAX_WIDTH = len(machine.machine.description)
                    ws['F'+ str(COUNT)] = machine.machine.measurement
                    ws['H'+ str(COUNT)] = data['ntz'][el[0].name]
                    ws['I'+ str(COUNT)] = data['total_ntz'][el[0].name]
                    COUNT += 1
            else:
                ws['D'+ str(COUNT)] = '...'
                COUNT += 1

            #for 16 string, materials
            ws['D'+str(COUNT)] = 'Материалы:'
            COUNT += 1
            if el[0].material_operation.all():
                for material in el[0].material_operation.all():
                    ws['D'+ str(COUNT)] = material.material.description
                    if len(material.material.description) >= MAX_WIDTH:
                        MAX_WIDTH = len(material.material.description)
                    ws['F'+ str(COUNT)] = material.material.measurement
                    ws['I'+ str(COUNT)] = material.count
                    COUNT += 1
            else:
                ws['D'+ str(COUNT)] = '...'
                COUNT += 1
            
            COUNT_OPERATION = COUNT
            COUNT += 1

        #for static and static excel
        COUNT_STATIC_END = COUNT_OPERATION
        set_border(ws, f'B{str(COUNT_STATIC)}:I{str(COUNT_STATIC_END-1)}', data['metadata'])

        end_document(ws=ws, count=COUNT_STATIC_END)    
        #witdh column
        ws.column_dimensions['D'].width = MAX_WIDTH
        return wb


class Form_3(Excel):
    def get_data(data):
    
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_3_example.xlsx')
        
        wb = load_workbook(file)
        ws = wb.active

        COUNT = 13
        START = 13

        #title
        ws['B6'] = data['object'].title
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'B6:C6')
        ws['B9'] = f"Измеритель ГЭСНм/ГЭСНп: {data['object'].main_measure.code}"


        #workmans string 13
        for i, el in enumerate(zip(data['workmans'].items(), data['measure_workmans'].values()),start=1):
            ws['A'+str(COUNT)] = i
            ws['B'+str(COUNT)] = el[0][0].name
            ws['C'+str(COUNT)] = el[0][1]
            ws['D'+str(COUNT)] = el[1]
            COUNT += 1

        #sum
        ws['B'+str(COUNT)] = "Вcего:"
        ws['C'+str(COUNT)] = data['sum_workmans']
        ws['D'+str(COUNT)] = data['sum_measure_workmans']
        COUNT += 1
        
        #k
        ws['B'+str(COUNT)] = f"K={data['object'].workman_k}"
        ws['C'+str(COUNT)] = data['k_sum_workmans']
        ws['D'+str(COUNT)] = data['k_sum_measure_workmans']
        COUNT += 1

        #avg
        ws['B'+str(COUNT)] = 'Средний разряд работы'
        ws['D'+str(COUNT)] = data['average_rank']
        set_border(ws, f'A{str(START)}:D{str(COUNT)}', data['metadata'])
        end_document(ws=ws, count=COUNT, start_row='a')
        
        return wb


class Form_4(Excel):
    def get_data(data):
        
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_4_example.xlsx')
        wb = load_workbook(file)
        ws = wb.active
        
        COUNT = 13
        START = 13
        MAX_WIDTH = 0

        #title
        ws['B6'] = data['object'].title
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'B6:E6')
        ws['A9'] = f"Измеритель ГЭСНм/ГЭСНп: {data['object'].main_measure.code}"

        #machines string 13
        for i, el in enumerate(zip(data['machines'].items(),data['is_driver'].values() ,data['measure_machines'].values()),start=1):
            ws['A'+str(COUNT)] = i
            ws['B'+str(COUNT)] = el[0][0].description
            if len(el[0][0].description) >= MAX_WIDTH:
                MAX_WIDTH = len(el[0][0].description)
            ws['C'+str(COUNT)] = el[0][1]
            ws['D'+str(COUNT)] = el[2]
            if el[1]:
                ws['E'+str(COUNT)] = el[0][1]
                ws['F'+str(COUNT)] = el[2]
            COUNT += 1
        ws.column_dimensions['B'].width = MAX_WIDTH

        #sum
        ws['B'+str(COUNT)] = 'Всего затраты труда машинистов'
        ws['F'+str(COUNT)] = data['sum_of_driver']

        set_border(ws, f'A{str(START)}:F{str(COUNT)}', data['metadata'])
        end_document(ws=ws, count=COUNT, start_row='a')

        return wb



class Form_5(Excel):
    def get_data(data):
    
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_5_example.xlsx')
        wb = load_workbook(file)
        ws = wb.active

        
        COUNT = 13
        START = 13
        MAX_WIDTH = 0

        #title
        ws['B6'] = data['object'].title
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'B6:D6')
        ws['A9'] = f"Измеритель ГЭСНм/ГЭСНп: {data['object'].main_measure.code}"

        #materials string 13
        for i, el in enumerate(zip(data['materials'].items(),data['measure'].values() ,data['material_measure'].values(), data['code'].values()),start=1):
            description_tbl = f"{el[3]} {el[0][0].description}"
            ws['A'+str(COUNT)] = i
            ws['B'+str(COUNT)] = description_tbl
            if len(description_tbl) >= MAX_WIDTH:
                MAX_WIDTH = len(description_tbl)
            ws['C'+str(COUNT)] = el[1]
            ws['D'+str(COUNT)] = el[0][1]
            ws['E'+str(COUNT)] = el[2]
            COUNT += 1
        ws.column_dimensions['B'].width = MAX_WIDTH
    
        set_border(ws, f'A{str(START)}:E{str(COUNT-1)}', data['metadata'])
        end_document(ws=ws, count=COUNT, start_row='a')
        return wb



class Form_6(Excel):
    def get_data(data):
    
        path = './core/static/excel_template'
        file = os.path.join(path, 'form_6_example.xlsx')
        wb = load_workbook(file)
        ws = wb.active
        ft = Font(name='Times New Roman')
        COUNT_OPER = 8
        MAX_WIDTH = 0

        ws['B3'] = f"Таблица ЭСН (ЭСНм, ЭСНр, ЭСНп) к {data['object'].code}"
        ws['B5'] = data['object'].title
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'B6:D6')

        for i, operation in enumerate(data['operations'], start=1):
            ws['B'+str(COUNT_OPER)] = f'{i} - {operation}'
            ws['B'+str(COUNT_OPER)].font = ft
            COUNT_OPER += 1

        COUNT_METER = COUNT_OPER + 2
        ws['B'+str(COUNT_METER)] = f"Измеритель ГЭСНм/ГЭСНп: {data['object'].main_measure.code}"
        COUNT_METER += 1


        START = COUNT_METER + 2
        COUNT_STATIC = COUNT_METER + 2
        ws['B'+str(COUNT_STATIC)] = 'Код ресурса'
        ws['C'+str(COUNT_STATIC)] = 'Наименование элемента затрат'
        ws['D'+str(COUNT_STATIC)] = 'Ед.изм.'
        ws['E'+str(COUNT_STATIC)] = 'Шифр'
        COUNT_STATIC += 1

        ws['B'+str(COUNT_STATIC)] = '1'
        ws['C'+str(COUNT_STATIC)] = 'Затраты труда рабочих-строителей'
        ws['D'+str(COUNT_STATIC)] = 'чел.-ч.'
        ws['E'+str(COUNT_STATIC)] = data['k_sum_measure_workmans']
        COUNT_STATIC += 1

        ws['B'+str(COUNT_STATIC)] = '1.1'
        ws['C'+str(COUNT_STATIC)] = 'Средний разряд работы'
        ws['E'+str(COUNT_STATIC)] = data['average_rank']
        COUNT_STATIC += 1

        ws['B'+str(COUNT_STATIC)] = '2'
        ws['C'+str(COUNT_STATIC)] = 'Затраты труда машинистов'
        ws['D'+str(COUNT_STATIC)] = 'чел.-ч.'
        ws['E'+str(COUNT_STATIC)] = data['sum_of_driver']
        COUNT_STATIC += 1

        ws['B'+str(COUNT_STATIC)] = '3'
        ws['C'+str(COUNT_STATIC)] = 'Машины и механизмы'
        ws['D'+str(COUNT_STATIC)] = 'маш.-ч.'
        COUNT_STATIC += 1

        #machine
        COUNT_MACHINE = COUNT_STATIC
        for key, value in data['measure_machines'].items():
            ws['C'+ str(COUNT_MACHINE)] = key.description
            if len(key.description) >= MAX_WIDTH:
                MAX_WIDTH = len(key.description)
            ws['D'+ str(COUNT_MACHINE)] = key.measurement
            ws['E'+ str(COUNT_MACHINE)] = value
            COUNT_MACHINE += 1
        ws.column_dimensions['C'].width = MAX_WIDTH

        COUNT_STATIC=COUNT_MACHINE
        ws['B'+str(COUNT_STATIC)] = '4'
        ws['C'+str(COUNT_STATIC)] = 'Материалы'
        COUNT_STATIC += 1


        #material
        COUNT_MATERIAL = COUNT_STATIC
        for key, value in data['material_measure'].items():
            ws['C'+ str(COUNT_MATERIAL)] = key.description
            ws['D'+ str(COUNT_MATERIAL)] = key.measurement
            ws['E'+ str(COUNT_MATERIAL)] = value
            COUNT_MATERIAL += 1


        set_border(ws, f'B{str(START)}:E{str(COUNT_MATERIAL - 1)}', data['metadata'])
        end_document(ws=ws, count=COUNT_MATERIAL)   
        return wb

from guide.machines_mechanisms import models as machines_mechanisms_models
from guide.classifier_kcp import models as classifier_kcp_models
from guide.snap_in_tool import models as snap_in_tool_models
from guide.tk_gesn import models
class ClassifierKcp(Excel):


    def get_data(ws):

        def fill_db():
            MODEL.Book.objects.create(
                            title='undefined',
                            description = 'undefined',
                        )
            
            MODEL.PartBook.objects.create(
                            title='undefined',
                            description='undefined',
                            book_id = 'undefined',
                        )

            MODEL.Chapter.objects.create(
                            title='undefined',
                            description='undefined',
                            partbook_id='undefined',
                        )

            MODEL.Group.objects.create(
                        title='undefined',
                        description='Материалы не вошедшие в справочник КСР',
                        chapter_id='undefined',
                    )



        ID_BOOK = 0
        ID_PARTBOOK = 0
        ID_CHAPTER = 0
        ID_GROUP = 0

        ACCEPT_PARTBOOK = False
        ACCEPT_CHAPTER = False
        ACCEPT_GROUP = False

        MISSED = 0
        MODEL = classifier_kcp_models
        MODEL.Book.objects.all().delete()
        MODEL.PartBook.objects.all().delete()
        MODEL.Chapter.objects.all().delete()
        MODEL.Group.objects.all().delete()
        MODEL.Resource.objects.all().delete()
        fill_db()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['J'], ws['K'], ws['L']):
            if MISSED < 3:
                MISSED += 1
                continue
            else:
                try:
                    book_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    book_title = ''
                book_model = MODEL.Book
                if el[1].value:
                    try:
                        ID_BOOK = book_model.objects.get(title=book_title).title
                    except book_model.DoesNotExist:
                        ID_BOOK = book_title
                        book_model.objects.create(
                            title=ID_BOOK,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                    ACCEPT_PARTBOOK = True
                    ACCEPT_CHAPTER = True
                    ACCEPT_GROUP = True
                    continue

                try:
                    partbook_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    partbook_title = ''
                partbook_model = MODEL.PartBook
                if partbook_title:
                    try:
                        ID_PARTBOOK = partbook_model.objects.get(title=partbook_title).title
                    except partbook_model.DoesNotExist:
                        ID_PARTBOOK = ID_BOOK+partbook_title
                        partbook_model.objects.create(
                            title=ID_PARTBOOK,
                            description = el[3].value if el[3].value else 'не задано',
                            book_id = ID_BOOK,
                        )
                    ACCEPT_PARTBOOK = False
                    ACCEPT_CHAPTER = True
                    ACCEPT_GROUP = True
                    continue
                elif ACCEPT_PARTBOOK:
                    ID_PARTBOOK = f"{ID_BOOK}"
                    partbook_model.objects.create(
                        title=ID_PARTBOOK,
                        description = el[3].value if el[3].value else 'не задано',
                        book_id = ID_BOOK,
                    )
                    ACCEPT_PARTBOOK = False


                try:
                    chapter_title = ''.join([i for i in el[4].value if i.isdigit()])
                except TypeError:
                    chapter_title = ''
                chapter_model = MODEL.Chapter
                if chapter_title:
                    try:
                        ID_CHAPTER = chapter_model.objects.get(title=chapter_title).title
                    except chapter_model.DoesNotExist:
                        ID_CHAPTER = ID_PARTBOOK+chapter_title
                        chapter_model.objects.create(
                            title=ID_CHAPTER,
                            description = el[5].value if el[5].value else 'не задано',
                            partbook_id = ID_PARTBOOK,
                        )
                    ACCEPT_PARTBOOK = False
                    ACCEPT_CHAPTER = False
                    ACCEPT_GROUP = True
                    continue

                elif ACCEPT_CHAPTER:
                    ID_CHAPTER=f"{ID_PARTBOOK}"
                    chapter_model.objects.create(
                        title=ID_CHAPTER,
                        description = el[5].value if el[5].value else 'не задано',
                        partbook_id = ID_PARTBOOK,
                    )
                    ACCEPT_CHAPTER = False

                try:
                    group_title = ''.join([i for i in el[6].value if i.isdigit()])
                except TypeError:
                    group_title = ''
                group_model = MODEL.Group
                if group_title:
                    try:
                        ID_GROUP = group_model.objects.get(title=group_title).title
                    except group_model.DoesNotExist:
                        ID_GROUP = ID_CHAPTER+group_title
                        group_model.objects.create(
                            title=ID_GROUP,
                            description = el[7].value if el[7].value else 'не задано',
                            chapter_id = ID_CHAPTER,
                        )
                    ACCEPT_PARTBOOK = False
                    ACCEPT_CHAPTER = False
                    ACCEPT_GROUP = False
                    continue
                elif ACCEPT_GROUP:
                    ID_GROUP=f"{ID_CHAPTER}"
                    group_model.objects.create(
                        title=ID_GROUP,
                        description = el[7].value if el[7].value else 'не задано',
                        chapter_id = ID_CHAPTER,
                    )
                    ACCEPT_GROUP = False


                resource_title = el[8].value
                resource_model = MODEL.Resource
                if el[8].value:
                    try:
                        resource_model.objects.get(title=resource_title).title
                    except resource_model.DoesNotExist:
                        resource_model.objects.create(
                            title=resource_title,
                            description = el[9].value if el[9].value else 'не задано',
                            measurement=el[10].value if el[10].value else 'не задано',
                            group_id = ID_GROUP,
                        )


class MachinesMechanisms(Excel):


    def get_data(ws):

        
        MISSED = 0
        ID_CHAPTER = 0
        ID_GROUP = 0

        ACCEPT_GROUP = False
        machines_mechanisms_models.Chapter.objects.all().delete()
        machines_mechanisms_models.Group.objects.all().delete()
        machines_mechanisms_models.Resource.objects.all().delete()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H']):
            if MISSED < 5:
                MISSED += 1
                continue
            else:
                try:
                    chapter_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    chapter_title = ''
                if chapter_title:
                    try:
                        ID_CHAPTER = machines_mechanisms_models.Chapter.objects.get(title=chapter_title).title
                    except machines_mechanisms_models.Chapter.DoesNotExist:
                        machines_mechanisms_models.Chapter.objects.create(
                            title=chapter_title,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                        ID_CHAPTER = chapter_title
                    ACCEPT_GROUP = True
                    continue

                    
                try:
                    group_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    group_title = ''
                if group_title:
                    try:
                        ID_GROUP = machines_mechanisms_models.Group.objects.get(title=group_title).title
                    except machines_mechanisms_models.Group.DoesNotExist:
                        machines_mechanisms_models.Group.objects.create(
                            title=group_title,
                            description = el[3].value if el[3].value else 'не задано',
                            chapter_id=ID_CHAPTER,
                        )
                        ID_GROUP = group_title
                    ACCEPT_GROUP = False
                    continue
                elif ACCEPT_GROUP:
                    machines_mechanisms_models.Group.objects.create(
                        title=f"{ID_CHAPTER}",
                         description = el[3].value if el[3].value else 'не задано',
                        chapter_id=ID_CHAPTER,
                    )
                    ACCEPT_GROUP = False


                resource_title = el[4].value
                if resource_title:
                    try:
                        machines_mechanisms_models.Resource.objects.get(title=resource_title).title
                    except machines_mechanisms_models.Resource.DoesNotExist:
                        machines_mechanisms_models.Resource.objects.create(
                            title=resource_title if resource_title else 'не задано',
                            description = el[5].value if el[5].value else 'не задано',
                            measurement = el[6].value if el[6].value else 'не задано',
                            group_id=ID_GROUP,
                        )
                    ACCEPT_GROUP = False
                        
                     



class SnapInTool(Excel):


    def get_data(ws):

        
        ID_CHAPTER = 0
        ID_GROUP = 0
        MISSED = 0
        ACCEPT_GROUP = False

        snap_in_tool_models.Chapter.objects.all().delete()
        snap_in_tool_models.Group.objects.all().delete()
        snap_in_tool_models.Resource.objects.all().delete()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H']):
            if MISSED < 4:
                MISSED += 1
                continue
            else:
                try:
                    chapter_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    chapter_title = ''
                chapter_model = snap_in_tool_models.Chapter
                if chapter_title:
                    try:
                        ID_CHAPTER = chapter_model.objects.get(title=chapter_title).title
                    except chapter_model.DoesNotExist:
                        ID_CHAPTER = chapter_title
                        chapter_model.objects.create(
                            title=ID_CHAPTER,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                    ACCEPT_GROUP = True
                    continue


                try:
                    group_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    group_title = ''
                group_model = snap_in_tool_models.Group
                if group_title:
                    try:
                        ID_GROUP = group_model.objects.get(title=group_title).title
                    except group_model.DoesNotExist:
                        ID_GROUP = ID_CHAPTER+group_title
                        group_model.objects.create(
                            title=ID_GROUP,
                            description = el[3].value if el[3].value else 'не задано',
                            chapter_id=ID_CHAPTER,
                        )
                    ACCEPT_GROUP = False
                    continue
                elif ACCEPT_GROUP:
                    ID_GROUP=f"{ID_CHAPTER}"
                    group_model.objects.create(
                        title=ID_GROUP,
                        description = el[3].value,
                        chapter_id=ID_CHAPTER,
                    )
                    ACCEPT_GROUP = False

                resource_title = el[4].value
                resource_model = snap_in_tool_models.Resource
                if resource_title:
                    try:
                        resource_model.objects.get(title=resource_title).title
                    except resource_model.DoesNotExist:
                        resource_model.objects.create(
                            title=resource_title,
                            description = el[5].value if el[5].value else 'не задано',
                            measurement = el[6].value if el[6].value else 'не задано',
                            group_id=ID_GROUP,
                        )
                    ACCEPT_GROUP = False

class TkMain(Excel):

    def get_data(ws):
       
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_SUBSECTION = 0
        ID_TABLESUBSECTION = 0
        
        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_SUBSECTION = False
        ACCEPT_TABLESUBSECTION = False
       
        models.TkMainCollection.objects.all().delete()
        models.TkMainDepartment.objects.all().delete()
        models.TkMainSection.objects.all().delete()
        models.TkMainSubsection.objects.all().delete()
        models.TkMainTableSubsection.objects.all().delete()
        models.TkMainTechCard.objects.all().delete()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K'], ws['L'], ws['M']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                try:
                    collection_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    collection_title = ''
                collection_model = models.TkMainCollection
                if collection_title:
                    try:
                        ID_COLLECTION = collection_model.objects.get(title=collection_title).title
                    except collection_model.DoesNotExist:
                        collection_model.objects.create(
                            title=collection_title,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                        ID_COLLECTION = collection_title
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                
                try:
                    department_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    department_title = ''
                department_model = models.TkMainDepartment
                if department_title:
                    try:
                        ID_DEPARTMENT = department_model.objects.get(title=department_title).title
                    except department_model.DoesNotExist:
                        ID_DEPARTMENT = ID_COLLECTION+department_title
                        department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT=f"{ID_COLLECTION}"
                    department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    

                try:
                    section_title = ''.join([i for i in el[4].value if i.isdigit()])
                except TypeError:
                    section_title = ''
                section_model = models.TkMainSection
                if section_title:
                    try:
                        ID_SECTION = section_model.objects.get(title=section_title).title
                    except section_model.DoesNotExist:
                        ID_SECTION = ID_DEPARTMENT+section_title
                        section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION=f"{ID_DEPARTMENT}"
                    section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                            )
                    ACCEPT_SECTION = False

                try:
                    subsection_title = ''.join([i for i in el[6].value if i.isdigit()])
                except TypeError:
                    subsection_title = ''
                subsection_model = models.TkMainSubsection
                if subsection_title:
                    try:
                        ID_SUBSECTION = subsection_model.objects.get(title=subsection_title).title
                    except subsection_model.DoesNotExist:
                        ID_SUBSECTION = ID_SECTION+subsection_title
                        subsection_model.objects.create(
                            title=ID_SUBSECTION,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SUBSECTION:
                    ID_SUBSECTION=f"{ID_SECTION}"
                    subsection_model.objects.create(
                            title=ID_SUBSECTION,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_SUBSECTION = False

                
                try:
                    table_subsection_title = ''.join([i for i in el[8].value if i.isdigit()])
                except TypeError:
                    table_subsection_title = ''
                table_subsection_model = models.TkMainTableSubsection
                if table_subsection_title:
                    try:
                        ID_TABLESUBSECTION = table_subsection_model.objects.get(title=table_subsection_title).title
                    except table_subsection_model.DoesNotExist:
                        ID_TABLESUBSECTION = ID_SUBSECTION+table_subsection_title
                        table_subsection_model.objects.create(
                            title=ID_TABLESUBSECTION,
                            description = el[9].value if el[9].value else 'не задано',
                            subsection_id = ID_SUBSECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = False
                    continue
                elif ACCEPT_TABLESUBSECTION:
                    ID_TABLESUBSECTION=f"{ID_SUBSECTION}"
                    table_subsection_model.objects.create(
                            title=ID_TABLESUBSECTION,
                            description = el[9].value if el[9].value else 'не задано',
                            subsection_id = ID_SUBSECTION
                            )
                    ACCEPT_TABLESUBSECTION = False
                    
                techcard_title = el[10].value
                techcard_model = models.TkMainTechCard      
                if techcard_title:
                    try:
                        techcard_model.objects.get(title=techcard_title).title
                    except techcard_model.DoesNotExist:
                        techcard_model.objects.create(
                            title=techcard_title,
                            description = el[11].value if el[11].value else 'не задано',
                            table_subsection_id = ID_TABLESUBSECTION
                        )
                    


class TkRemoteConstruction(Excel):

    def get_data(ws):
        
        
        MISSED = 4
        ID_COLLECTION = None
        ID_DEPARTMENT = None
        ID_SECTION = None
        ID_SUBSECTION = None

        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_SUBSECTION = False
        models.TkRemoteConstructionCollection.objects.all().delete()
        models.TkRemoteConstructionDepartment.objects.all().delete()
        models.TkRemoteConstructionSection.objects.all().delete()
        models.TkRemoteConstructionSubsection.objects.all().delete()
        models.TkRemoteConstructionTechCard.objects.all().delete()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                try:
                    collection_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    collection_title = ''
                collection_model = models.TkRemoteConstructionCollection
                if collection_title:
                    try:
                        ID_COLLECTION = collection_model.objects.get(title=collection_title).title
                    except collection_model.DoesNotExist:
                        collection_model.objects.create(
                            title=collection_title,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                        ID_COLLECTION = collection_title
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    continue
                

                try:
                    department_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    department_title = ''
                department_model = models.TkRemoteConstructionDepartment
                if department_title:
                    try:
                        ID_DEPARTMENT = department_model.objects.get(title=department_title).title
                    except department_model.DoesNotExist:
                        ID_DEPARTMENT = ID_COLLECTION+department_title
                        department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT=f"{ID_COLLECTION}"
                    department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    

                try:
                    section_title = ''.join([i for i in el[4].value if i.isdigit()])
                except TypeError:
                    section_title = ''
                section_model = models.TkRemoteConstructionSection
                if section_title:
                    try:
                        ID_SECTION = section_model.objects.get(title=section_title).title
                    except section_model.DoesNotExist:
                        ID_SECTION = ID_DEPARTMENT+section_title
                        section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = True
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION=f"{ID_DEPARTMENT}"
                    section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                            )
                    ACCEPT_SECTION = False


                try:
                    subsection_title = ''.join([i for i in el[6].value if i.isdigit()])
                except TypeError:
                    subsection_title = ''
                subsection_model = models.TkRemoteConstructionSubsection
                if subsection_title:
                    try:
                        ID_SUBSECTION = subsection_model.objects.get(title=subsection_title).title
                    except subsection_model.DoesNotExist:
                        ID_SUBSECTION = ID_SECTION+subsection_title
                        subsection_model.objects.create(
                            title=ID_SUBSECTION,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    continue
                elif ACCEPT_SUBSECTION:
                    ID_SUBSECTION=f"{ID_SECTION}"
                    subsection_model.objects.create(
                            title=ID_SUBSECTION,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_SUBSECTION = False
                    
                        
                techcard_title = el[8].value
                techcard_model = models.TkRemoteConstructionTechCard      
                if techcard_title:
                    try:
                        techcard_model.objects.get(title=techcard_title).title
                    except techcard_model.DoesNotExist:
                        techcard_model.objects.create(
                            title=techcard_title,
                            description = el[9].value if el[9].value else 'не задано',
                            subsection_id = ID_SUBSECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False


class TkInstallationEquipment(Excel):

    def get_data(ws):
        
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_TABLE = 0

        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_TABLE = False

        models.TkInstallationEquipmentCollection.objects.all().delete()
        models.TkInstallationEquipmentDepartment.objects.all().delete()
        models.TkInstallationEquipmentSection.objects.all().delete()
        models.TkInstallationEquipmentTable.objects.all().delete()
        models.TkInstallationEquipmentTechCard.objects.all().delete()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                try:
                    collection_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    collection_title = ''
                collection_model = models.TkInstallationEquipmentCollection
                if collection_title:
                    try:
                        ID_COLLECTION = collection_model.objects.get(title=collection_title).title
                    except collection_model.DoesNotExist:
                        ID_COLLECTION = collection_title
                        collection_model.objects.create(
                            title=ID_COLLECTION,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                        ID_COLLECTION = collection_title
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_TABLE = True
                    continue
                

                try:
                    department_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    department_title = ''
                department_model = models.TkInstallationEquipmentDepartment
                if department_title:
                    try:
                        ID_DEPARTMENT = department_model.objects.get(title=department_title).title
                    except department_model.DoesNotExist:
                        ID_DEPARTMENT = ID_COLLECTION+department_title
                        department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_TABLE = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT=f"{ID_COLLECTION}"
                    department_model.objects.create(
                        title=ID_DEPARTMENT,
                        description = el[3].value if el[3].value else 'не задано',
                        collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    


                try:
                    section_title = ''.join([i for i in el[4].value if i.isdigit()])
                except TypeError:
                    section_title = ''
                section_model = models.TkInstallationEquipmentSection
                if section_title:
                    try:
                        ID_SECTION = section_model.objects.get(title=section_title).title
                    except section_model.DoesNotExist:
                        ID_SECTION = ID_DEPARTMENT+section_title
                        section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_TABLE = True
                    continue

                elif ACCEPT_SECTION:
                    ID_SECTION=f"{ID_DEPARTMENT}"
                    section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                        )
                    ACCEPT_SECTION = False

                try:
                    table_title = ''.join([i for i in el[6].value if i.isdigit()])
                except TypeError:
                    table_title = ''
                table_model = models.TkInstallationEquipmentTable
                if table_title:
                    try:
                        ID_TABLE = table_model.objects.get(title=table_title).title
                    except table_model.DoesNotExist:
                        ID_TABLE = ID_SECTION+table_title
                        table_model.objects.create(
                            title=ID_TABLE,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_TABLE = False
                    continue
                elif ACCEPT_TABLE:
                    ID_TABLE = f"{ID_SECTION}"
                    table_model.objects.create(
                        title=ID_TABLE,
                        description = el[7].value if el[7].value else 'не задано',
                        section_id = ID_SECTION
                    )
                    ACCEPT_TABLE = False
                    
                        
                techcard_title = el[8].value
                techcard_model = models.TkInstallationEquipmentTechCard      
                if techcard_title:
                    try:
                        techcard_model.objects.get(title=techcard_title).title
                    except techcard_model.DoesNotExist:
                        techcard_model.objects.create(
                            title=techcard_title,
                            description = el[9].value if el[9].value else 'не задано',
                            table_id = ID_TABLE
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_TABLE = False


class TkPre_commissioning(Excel):

    def get_data(ws):
        
        MISSED = 4
        ID_COLLECTION = 0
        ID_DEPARTMENT = 0
        ID_SECTION = 0
        ID_SUBSECTION = 0
        ID_TABLESUBSECTION = 0

        ACCEPT_DEPARTMENT = False
        ACCEPT_SECTION = False
        ACCEPT_SUBSECTION = False
        ACCEPT_TABLESUBSECTION = False

        models.TkPre_commissioningCollection.objects.all().delete()
        models.TkPre_commissioningDepartment.objects.all().delete()
        models.TkPre_commissioningSection.objects.all().delete()
        models.TkPre_commissioningSubsection.objects.all().delete()
        models.TkPre_commissioningTableSubsection.objects.all().delete()
        models.TkPre_commissioningTechCard.objects.all().delete()
        for el in zip(ws['B'], ws['C'], ws['D'], ws['E'],ws['F'] ,ws['G'], ws['H'], ws['I'], ws['j'], ws['K'], ws['L'], ws['M']):
            if MISSED != 0:
                MISSED -= 1
                continue
            else:
                try:
                    collection_title = ''.join([i for i in el[0].value if i.isdigit()])
                except TypeError:
                    collection_title = ''
                collection_model = models.TkPre_commissioningCollection
                if el[0].value:
                    try:
                        ID_COLLECTION = collection_model.objects.get(title=collection_title).title
                    except collection_model.DoesNotExist:
                        collection_model.objects.create(
                            title=collection_title,
                            description = el[1].value if el[1].value else 'не задано',
                        )
                        ID_COLLECTION = collection_title
                    ACCEPT_DEPARTMENT = True
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                
                try:
                    department_title = ''.join([i for i in el[2].value if i.isdigit()])
                except TypeError:
                    department_title = ''
                department_model = models.TkPre_commissioningDepartment
                if el[2].value:
                    try:
                        ID_DEPARTMENT = department_model.objects.get(title=department_title).title
                    except department_model.DoesNotExist:
                        ID_DEPARTMENT = ID_COLLECTION+department_title
                        department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = True
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_DEPARTMENT:
                    ID_DEPARTMENT=f"{ID_COLLECTION}"
                    department_model.objects.create(
                            title=ID_DEPARTMENT,
                            description = el[3].value if el[3].value else 'не задано',
                            collection_id = ID_COLLECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    

                try:
                    section_title = ''.join([i for i in el[4].value if i.isdigit()])
                except TypeError:
                    section_title = ''
                section_model = models.TkPre_commissioningSection
                if el[4].value:
                    try:
                        ID_SECTION = section_model.objects.get(title=section_title).title
                    except section_model.DoesNotExist:
                        ID_SECTION = ID_DEPARTMENT+section_title
                        section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = True
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SECTION:
                    ID_SECTION=f"{ID_DEPARTMENT}"
                    section_model.objects.create(
                            title=ID_SECTION,
                            description = el[5].value if el[5].value else 'не задано',
                            department_id = ID_DEPARTMENT
                            )
                    ACCEPT_SECTION = False

                try:
                    subsection_title = ''.join([i for i in el[6].value if i.isdigit()])
                except TypeError:
                    subsection_title = ''
                subsection_model = models.TkPre_commissioningSubsection
                if el[6].value:
                    try:
                        ID_SUBSECTION = subsection_model.objects.get(title=subsection_title).title
                    except subsection_model.DoesNotExist:
                        ID_SUBSECTION = ID_SECTION+subsection_title
                        subsection_model.objects.create(
                            title=ID_SUBSECTION,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = True
                    continue
                elif ACCEPT_SUBSECTION:
                    ID_SUBSECTION=f"{ID_SECTION}"
                    subsection_model.objects.create(
                            title=ID_SUBSECTION,
                            description = el[7].value if el[7].value else 'не задано',
                            section_id = ID_SECTION
                        )
                    ACCEPT_SUBSECTION = False

                
                try:
                    table_subsection_title = ''.join([i for i in el[8].value if i.isdigit()])
                except TypeError:
                    table_subsection_title = ''
                table_subsection_model = models.TkPre_commissioningTableSubsection
                if el[8].value:
                    try:
                        ID_TABLESUBSECTION = table_subsection_model.objects.get(title=table_subsection_title).title
                    except table_subsection_model.DoesNotExist:
                        ID_TABLESUBSECTION = ID_SUBSECTION+table_subsection_title
                        table_subsection_model.objects.create(
                            title=ID_TABLESUBSECTION,
                            description = el[9].value if el[9].value else 'не задано',
                            subsection_id = ID_SUBSECTION
                        )
                    ACCEPT_DEPARTMENT = False
                    ACCEPT_SECTION = False
                    ACCEPT_SUBSECTION = False
                    ACCEPT_TABLESUBSECTION = False
                    continue
                elif ACCEPT_TABLESUBSECTION:
                    ID_TABLESUBSECTION=f"{ID_SUBSECTION}"
                    table_subsection_model.objects.create(
                            title=ID_TABLESUBSECTION,
                            description = el[9].value if el[9].value else 'не задано',
                            subsection_id = ID_SUBSECTION
                            )
                    ACCEPT_TABLESUBSECTION = False
                    
                techcard_title = el[10].value
                techcard_model = models.TkPre_commissioningTechCard      
                if techcard_title:
                    try:
                        techcard_model.objects.get(title=techcard_title).title
                    except techcard_model.DoesNotExist:
                        techcard_model.objects.create(
                            title=techcard_title,
                            description = el[11].value if el[11].value else 'не задано',
                            table_subsection_id = ID_TABLESUBSECTION
                        )
